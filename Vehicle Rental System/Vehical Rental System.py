
import openpyxl
import openpyxl.workbook
import random as r
import string as s



# Dictionary of the Vehicles available for rent

Vehicle = {"Car": [{"ID": "C1", "Milage": "200", "Rent": " 300/hour", "Quantity Available": "7 units"},
                   {"ID": "C2", "Milage": "300", "Rent": " 500/hour", "Quantity Available": "9 units"}, 
                   {"ID": "C3", "Milage": "350", "Rent": " 600/hour", "Quantity Available": "11 units"},
                   {"ID": "C3", "Milage": "350", "Rent": " 600/hour", "Quantity Available": "17 units"}, 
                   {"ID": "C3", "Milage": "250", "Rent": " 400/hour", "Quantity Available": "28 units"}],

            "Cycle": [{"ID": "Cy1", "Milage": 'N/A', "Rent":"120/hour", "Quantity Available": "9 units"},
                      {"ID": "Cy1", "Milage": 'N/A', "Rent":"150/hour", "Quantity Available": "1 units"},
                      {"ID": "Cy1", "Milage": 'N/A', "Rent":"100/hour", "Quantity Available": "8 units"},
                      {"ID": "Cy1", "Milage": 'N/A', "Rent":"110/hour", "Quantity Available": "5 units"},
                      {"ID": "Cy1", "Milage": 'N/A', "Rent":"140/hour", "Quantity Available": "3 units"}],

            "Bike": [{"ID": "B1", "Milage": "120", "Rent": "240/hour", "Quantity Available": "6 units"},
                    {"ID": "B1", "Milage": "120", "Rent": "240/hour", "Quantity Available": "5 units"},
                    {"ID": "B1", "Milage": "120", "Rent": "240/hour", "Quantity Available": "4 units"},
                    {"ID": "B1", "Milage": "120", "Rent": "240/hour", "Quantity Available": "3 units"},
                    {"ID": "B1", "Milage": "120", "Rent": "240/hour", "Quantity Available": "2 units"}]
            }

for i,j in Vehicle.items():
    print(f"{i}: ")
    for k in j:
        print(f"\nVehicle ID: {k['ID']}\nMilage of Vehicle: {k['Milage']}\nRent/hour: {k['Rent']}\nQuantity Available: {k['Quantity Available']}")


                # Miscellenous Functions

# For Email check

def check_email():
    while True:
        E = input("Enter your full E-mail address :   ") 
        if "@" not in E or "." not in E:
            print("Please insert full email address inclusive of '@' and '.' ")
            continue 
        else:
            return E

# For contact number check

def check_contact():
    while True:
        C = input("Enter your contact number:   ")
        if len(C) == 10 and C.isdigit():
            return C
        else:
            print("Plese enter valid 10 digit contact number")
            continue



# Encryption of the Password

def encryp():
    while True:
        P = input("Enter your password:   ")
        if len(P) > 8:
            password = P
            chars = s.punctuation + s.ascii_lowercase + s.digits + s.ascii_uppercase
            chars = list(chars)

            r.seed(11)
            chars_copy = chars.copy()
            r.shuffle(chars_copy)
            pass_en = ""

            for i in password:
                index_ = chars.index(i)
                pass_en += chars_copy[index_]
            return pass_en
        else:
            print("Password should be more than 8 characters")
            continue
    


# Decryption of the Password

def decryp(): 
    # E_Pass = encryp()
    while True:
        E_Pass = input("Enter your password:   ")
        if len(E_Pass) > 8:
            chars = s.punctuation + s.ascii_lowercase + s.digits + s.ascii_uppercase
            chars = list(chars)

            r.seed(11)
            chars_copy = chars.copy()
            r.shuffle(chars_copy)
            pass_de = ""

            for i in E_Pass:
                index_ = chars.index(i)
                pass_de += chars_copy[index_]
            return pass_de
        else:
            print("Password should be more than 8 characters")
            continue









# Vehicle Class

class Vehicle:
    def __init__(self):
        pass

    def check_availability(self):
        if self.availabilty is True:
            return "Available"
        else:
            return "Not Available"
        
    def update_rate(self):
        new_rate = int(input("Enter the new rate"))
        self.rate = new_rate

    def list_vehicle(self):
        for i,j in Vehicle.items():
            print(f"{i}: ")
            for k in j:
                print(f"\nVehicle ID: {k['ID']}\nMilage of Vehicle: {k['Milage']}\nRent/hour: {k['Rent']}")

        


# Customer Class

class Customer:
    def __init__(self):
        self.account = None
        self.wb = openpyxl.load_workbook("VRS.xlsx")
        self.ws = self.wb.active

    def customer_menu(self):
        print("WELCOME TO YOUR VEHICLE RENTAL SERVICE".center(58, "-"))
        print("\n1. Create Account")
        print("2. Login")
        print("3. Exit")
        try:
            choice = input("Enter your choice:  ")
            if choice == "1":
                self.register()
            elif choice == "2":
                self.login()
            elif choice == "3":
                print("Thank you for using our service")
                exit()
            else:
                print("Input valid choice.")
        except ValueError:
                print("Invalid choice. Please try again.")

        
    # Registeration for a customer

    def register(self):
        F_Name = input("Enter your Full Name:  ")
        Email = check_email()
        Phone = check_contact()
        Password = encryp()
        # for i in range(1,100):
        i=1
        while self.ws.cell(row= i+1, column= 1).value is not None:
            i+=1
        self.ws.cell(row= i+1, column= 1).value = F_Name
        self.ws.cell(row= i+1, column= 2).value = Email
        self.ws.cell(row= i+1, column = 3).value = Phone
        self.ws.cell(row= i+1, column= 4).value = Password
        print("Account created successfully!!")
        self.wb.save("VRS.xlsx")

        # self.account = 
        print(f"You have been registed as: {F_Name} ")
            
        # return F_Name, Email, Phone, Password
    


        # Loging in for a customer

    def login(self):
        Email = input("Enter your Email:    ")
        Password = decryp()
        for i in range(2, self.ws.max_row + 1):
            stored_email = self.ws.cell(row=i, column=2).value
            stored_password = self.ws.cell(row=i, column=4).value
            name = self.ws.cell(row=i, column=1).value
            if stored_email == Email and stored_password == Password:
                print("Login successful!!")
                print(f"You have been registed as: {name} ")
                break

                # return Email, Password
        else:
            print("Invalid Email or Password.")




C1 = Customer()
print(C1.customer_menu())




